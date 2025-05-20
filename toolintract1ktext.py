import time
import logging
from openpyxl import load_workbook
import random
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium import webdriver
import threading
import queue
import os
import shutil
from selenium.webdriver.common.action_chains import ActionChains
import random
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import requests
from selenium.webdriver.common.action_chains import ActionChains
import re
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
import string
from selenium.webdriver.support.ui import WebDriverWait
import os
from selenium import webdriver
import time
import json
from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import subprocess
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.support.ui import Select
api_key = "CAP-B064D21B508CAED66557292EF0ACBD43"  # your api key of capsolver
site_key = "6LcioXgnAAAAAKAAs-0CkUiqrHdG-HRGgu3EYLEZ"  # site key of your target site
site_url = "https://gcp-api.intract.io"  # page url of your target site
 
 
def capsolver():
    payload = {
        "clientKey": api_key,
        "task": {
            "type": 'ReCaptchaV2TaskProxyLess',
            "websiteKey": site_key,
            "websiteURL": site_url
        }
    }
    res = requests.post("https://api.capsolver.com/createTask", json=payload)
    resp = res.json()
    task_id = resp.get("taskId")
    if not task_id:
        print("Failed to create task:", res.text)
        return
    print(f"Got taskId: {task_id} / Getting result...")
 
    while True:
        time.sleep(3)  # delay
        payload = {"clientKey": api_key, "taskId": task_id}
        res = requests.post("https://api.capsolver.com/getTaskResult", json=payload)
        resp = res.json()
        status = resp.get("status")
        if status == "ready":
            return resp.get("solution", {}).get('gRecaptchaResponse')
        if status == "failed" or resp.get("errorId"):
            print("Solve failed! response:", res.text)
            return
def worker(task_queue, counter_queue):
    while True:
        try:
            item = task_queue.get(timeout=60)  # Chờ tối đa 60 giây
        except queue.Empty:
            logging.warning("Worker timeout do không có task mới")
            break
        
        if item is None:
            break
        
        tk, mk, fa, profile, diachivi, keyokx, keyok1222 = item
        counter = counter_queue.get()

        try:
            start(tk, mk, fa, counter, profile, diachivi, keyokx, keyok1222)
        except Exception as e:
            logging.error(f"Lỗi trong worker: {e}")
        finally:
            task_queue.task_done()
            counter_queue.put(counter)
        
        time.sleep(10)


def load_extension_page(driver, url, max_load_time=10, retries=100):
    attempt = 0
    while attempt < retries:
        try:
            # Thiết lập thời gian chờ trang tải xong
            driver.set_page_load_timeout(max_load_time)
            driver.get(url)
            
            # Sử dụng WebDriverWait để chờ trang tải xong
            WebDriverWait(driver, max_load_time).until(
                EC.presence_of_element_located((By.TAG_NAME, 'body'))
            )
            print(f"Trang web {url} đã tải xong.")
            break  # Thoát khỏi vòng lặp nếu trang tải xong
        except (TimeoutException, WebDriverException) as e:
            print(f"Thử lại {attempt + 1} do trang web không tải xong trong {max_load_time} giây hoặc xảy ra lỗi: {e}")
            attempt += 1
            time.sleep(2)  # Chờ một chút trước khi thử lại
    else:
        print(f"Trang web {url} không thể tải xong sau {retries} lần thử.")
def xacminhmetamsk(driver):
    
    time.sleep(5)
    
    while True:
        try:
            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[-1])
            try:
                driver.implicitly_wait(2)
                driver.find_element(By.XPATH,"//button[text()='Got it']").click()
            except:
                print("không cần")
            time.sleep(1)
            script = "document.querySelector('.button.btn--rounded.btn-primary').click();"

            # Execute the script
            driver.execute_script(script)
            time.sleep(1)

        except:
            print("đã kết nối thành công")
            break


    window_handles = driver.window_handles
    driver.switch_to.window(window_handles[0])    
def logmetmaks(driver,keyokx):
    tk2 = keyokx
    tai = "1"
    nhap123 = "1"
    # chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html#onboarding/welcome
    if nhap123 == "2":
        link1231 = "chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/popup.html"
        link12312 = 'chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/popup.html#unlock'
        link123123 = 'chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html#onboarding/unlock'
    else:
        link1231 = "chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html#onboarding/welcome"
        link12312 = "chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html#unlock"
    try:
        window_handles = driver.window_handles
        driver.switch_to.window(window_handles[0])

    except:
        print("pm")
    load_extension_page(driver,link1231)
    time.sleep(0.1)
    try:
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//p[text()='Mantle Testnet']").click()
        time.sleep(1)
        try:
            driver.find_element(By.XPATH,"//span[text()='Ethereum Mainnet']").click()
        except:
            try:
                driver.find_element(By.XPATH,"//span[text()='BNB Smart Chain']").click()
            except:
                print("ok rooif")
        time.sleep(1)
    except:
        print("ok")
    time.sleep(2)
    if link12312  in driver.current_url or link12312 in driver.current_url:
        if tai == "1":
            driver.implicitly_wait(500)
            driver.find_element(By.XPATH,'//input[@data-testid="unlock-password"]').send_keys("A12345678a")
        else:
            load_extension_page(driver,"chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html#restore-vault")
            time.sleep(1)
            driver.implicitly_wait(500)
            words = tk2.split()

            # Check if the number of words is equal to 12
            if len(words) != 12:
                print("The number of words is not equal to 12")
            else:
                # Assign each word to a separate variable
                var1, var2, var3, var4, var5, var6, var7, var8, var9, var10, var11, var12 = words
                
                # Print the values of the variables
                print(var1)
                print(var2)
                print(var3)
                print(var4)
                print(var5)
                print(var6)
                print(var7)
                print(var8)
                print(var9)
                print(var10)
                print(var11)
                print(var12)

            driver.implicitly_wait(10)
            driver.find_element(By.ID,"import-srp__srp-word-0").send_keys(var1)
            driver.implicitly_wait(10)
            driver.find_element(By.ID,"import-srp__srp-word-1").send_keys(var2)
            driver.implicitly_wait(10)
            driver.find_element(By.ID,"import-srp__srp-word-2").send_keys(var3)
            driver.implicitly_wait(10)
            driver.find_element(By.ID,"import-srp__srp-word-3").send_keys(var4)
            driver.implicitly_wait(10)
            driver.find_element(By.ID,"import-srp__srp-word-4").send_keys(var5)
            driver.implicitly_wait(10)
            driver.find_element(By.ID,"import-srp__srp-word-5").send_keys(var6)
            driver.implicitly_wait(10)
            driver.find_element(By.ID,"import-srp__srp-word-6").send_keys(var7)
            driver.implicitly_wait(10)
            driver.find_element(By.ID,"import-srp__srp-word-7").send_keys(var8)
            driver.implicitly_wait(10)
            driver.find_element(By.ID,"import-srp__srp-word-8").send_keys(var9)
            driver.implicitly_wait(10)
            driver.find_element(By.ID,"import-srp__srp-word-9").send_keys(var10)
            driver.implicitly_wait(10)
            driver.find_element(By.ID,"import-srp__srp-word-10").send_keys(var11)
            driver.implicitly_wait(10)
            driver.find_element(By.ID,"import-srp__srp-word-11").send_keys(var12)
            driver.implicitly_wait(10)
            driver.find_element(By.ID,'password').send_keys("A12345678a")
            time.sleep(1)
            driver.implicitly_wait(10)
            driver.find_element(By.ID,'confirm-password').send_keys("A12345678a")
            time.sleep(1)
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,'//button[@data-testid="create-new-vault-submit-button"]').click()
        time.sleep(0.1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//button[text()='Unlock']").click()
        try:
            time.sleep(5)
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,"//button[text()='Got it!']").click()
            window_handles = driver.window_handles

            
        except:
            window_handles = driver.window_handles

            
    else:
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"onboarding__terms-checkbox").click()
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//button[text()='Import an existing wallet']").click()
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//button[text()='I agree']").click()
        time.sleep(1)
        words = tk2.split()

        # Check if the number of words is equal to 12
        if len(words) != 12:
            print("The number of words is not equal to 12")
        else:
            # Assign each word to a separate variable
            var1, var2, var3, var4, var5, var6, var7, var8, var9, var10, var11, var12 = words
            
            # Print the values of the variables
            print(var1)
            print(var2)
            print(var3)
            print(var4)
            print(var5)
            print(var6)
            print(var7)
            print(var8)
            print(var9)
            print(var10)
            print(var11)
            print(var12)

        driver.implicitly_wait(10)
        driver.find_element(By.ID,"import-srp__srp-word-0").send_keys(var1)
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"import-srp__srp-word-1").send_keys(var2)
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"import-srp__srp-word-2").send_keys(var3)
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"import-srp__srp-word-3").send_keys(var4)
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"import-srp__srp-word-4").send_keys(var5)
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"import-srp__srp-word-5").send_keys(var6)
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"import-srp__srp-word-6").send_keys(var7)
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"import-srp__srp-word-7").send_keys(var8)
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"import-srp__srp-word-8").send_keys(var9)
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"import-srp__srp-word-9").send_keys(var10)
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"import-srp__srp-word-10").send_keys(var11)
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"import-srp__srp-word-11").send_keys(var12)
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//button[text()='Confirm Secret Recovery Phrase']").click()
        time.sleep(2)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,'//input[@data-testid="create-password-new"]').send_keys("A12345678a")
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,'//input[@data-testid="create-password-confirm"]').send_keys("A12345678a")
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//input[@type='checkbox']").click()
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//button[text()='Import my wallet']").click()
        time.sleep(10)
        driver.implicitly_wait(10)
        while True:
            if "chrome-extension://nkbihfbeogaeaoehlefnkodbefgpgknn/home.html#onboarding/completion" in driver.current_url:
                break
            else:
                time.sleep(5)
        driver.execute_script('document.querySelector("#app-content > div > div.mm-box.main-container-wrapper > div > div > div > div.box.creation-successful__actions.box--margin-top-6.box--flex-direction-row > button").click()')
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//button[text()='Next']").click()
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//button[text()='Done']").click()
        time.sleep(5)
        # driver.get("https://www.google.com/")
        # time.sleep(5)
    driver.get("https://www.google.com/")
    time.sleep(1)
    window_handles = driver.window_handles
    driver.switch_to.window(window_handles[0])

def ketnoilaicopopup1(driver):

    start_time = time.time()  # Record the start time
    
    while True:
        try:
            if time.time() - start_time > 120:
                print("Timeout reached. Exiting loop.")
                driver.quit()
                break
            # Lưu lại tất cả các cửa sổ đang mở
            window_handles = driver.window_handles

            # Duyệt qua từng cửa sổ
            for handle in window_handles:
                driver.switch_to.window(handle)
                # Kiểm tra URL của tab
                if driver.current_url.startswith("chrome-extension://"):
                    break  # Nếu đúng tab cần tìm, dừng vòng lặp

            driver.implicitly_wait(10)
            button = driver.find_element(By.CSS_SELECTOR, 'button.okui-btn.btn-lg.btn-fill-highlight.mobile._action-button_1ntoe_1')
            time.sleep(2)
            button.click()
            
            break
            

        except:
            print("loix")
            
                


    # Đảm bảo quay lại cửa sổ chính
    try:
        window_handles = driver.window_handles
        driver.switch_to.window(window_handles[0])
    except Exception as e:
        print(f"lỗi: {e}")

# Gọi hàm với driver (đảm bảo bạn đã khởi tạo và cấu hình driver thích hợp)
# ketnoilaicopopup1(driver)

    
def themtwiter(driver,tk,mk,fa):
    # driver.get(str(link1221))
    while True:
        try:
            driver.implicitly_wait(5)
            driver.find_element(By.XPATH,"//div[@class='social-account-link-text' and contains(text(), 'Connect Twitter Account')]").click()
            break
        except:
            scroll_down_pixels = 10

            # Execute JavaScript code to scroll down
            driver.execute_script(f"window.scrollBy(0, {scroll_down_pixels});")
    time.sleep(2)
    driver.find_element(By.XPATH,"//button[@class='tc-tweet-button']").click()
    time.sleep(2)
    linknoidung = "https://x.com/intent/post"
    post(driver,tk,mk,fa,linknoidung)
    time.sleep(2)
    
    login_twitter(driver,tk,mk,fa)
    driver.get("https://x.com/")
    
    driver.implicitly_wait(30)
    driver.find_element(By.XPATH,"//a[@data-testid='AppTabBar_Profile_Link']").click()
    time.sleep(5)
    link = driver.current_url
    if link == "https://x.com/account/access":
        kiemtracapcha(driver)
        driver.get("https://x.com")
        driver.implicitly_wait(30)
        driver.find_element(By.XPATH,"//a[@data-testid='AppTabBar_Profile_Link']").click()
    else:
        print("tiep tuc")
    if "Something went wrong. Try reloading." in driver.page_source:
        linl666 = driver.current_url
        login_twitter1(driver,tk,mk,fa)
        driver.get(str(linl666))
    driver.implicitly_wait(10)
    
    max_time = 30  # Thời gian tối đa (giây)
    start_time = time.time()

    # Vòng lặp while
    while True:
        try:
            # Tìm và nhấp vào phần tử
            driver.find_element(By.XPATH, "//a[contains(@aria-label, 'View post analytics')]").click()
            break
        except NoSuchElementException:
            # Nếu không tìm thấy phần tử, cuộn xuống
            scroll_down_pixels = 10
            driver.execute_script(f"window.scrollBy(0, {scroll_down_pixels});")
            
            # Kiểm tra xem đã vượt quá thời gian tối đa chưa
            elapsed_time = time.time() - start_time
            if elapsed_time >= max_time:
                print("Quá 30s.")
                driver.quit()
                break
    time.sleep(2)
    linkanitic = driver.current_url
    # Tách chuỗi URL bằng dấu "/"
    parts = linkanitic.split("/")

    # Loại bỏ phần tử cuối cùng, tức là "/analytics"
    new_link = "/".join(parts[:-1])

    print(new_link)  # In ra kết quả
    window_handles = driver.window_handles
    driver.switch_to.window(window_handles[0])
    driver.implicitly_wait(10)
    driver.find_element(By.XPATH,"//input[@placeholder='Paste link here']").send_keys(str(new_link))
    time.sleep(0.1)
    while True:
        try:
            driver.find_element(By.XPATH,"//span[text()='Verify']").click()
            break
        except:
            scroll_down_pixels = 10

            # Execute JavaScript code to scroll down
            driver.execute_script(f"window.scrollBy(0, {scroll_down_pixels});")
    time.sleep(2)
    try:
        driver.execute_script('document.querySelector("#app > div.v-dialog__content.v-dialog__content--active > div > div > div > div.footer.flex-align-center.flex-justify-end > div > div:nth-child(2) > button > span").click()')
    except:
        print("không cần")
def themlaitwiter(driver,tk,mk,fa):
    # driver.get(str(link1221))
    
    time.sleep(2)
    driver.find_element(By.XPATH,"//button[@class='tc-tweet-button']").click()
    time.sleep(2)
    linknoidung = "https://x.com/intent/post"
    post(driver,tk,mk,fa,linknoidung)
    time.sleep(2)
    login_twitter(driver,tk,mk,fa)
    time.sleep(2)
    driver.get("https://x.com/")
    
    driver.implicitly_wait(30)
    driver.find_element(By.XPATH,"//a[@data-testid='AppTabBar_Profile_Link']").click()
    time.sleep(5)
    if "Something went wrong. Try reloading." in driver.page_source:
        linl666 = driver.current_url
        login_twitter1(driver,tk,mk,fa)
        driver.get(str(linl666))
    driver.implicitly_wait(10)
    
    max_time = 30  # Thời gian tối đa (giây)
    start_time = time.time()

    # Vòng lặp while
    while True:
        try:
            # Tìm và nhấp vào phần tử
            driver.find_element(By.XPATH, "//a[contains(@aria-label, 'View post analytics')]").click()
            break
        except NoSuchElementException:
            # Nếu không tìm thấy phần tử, cuộn xuống
            scroll_down_pixels = 10
            driver.execute_script(f"window.scrollBy(0, {scroll_down_pixels});")
            
            # Kiểm tra xem đã vượt quá thời gian tối đa chưa
            elapsed_time = time.time() - start_time
            if elapsed_time >= max_time:
                print("Quá 30s.")
                driver.quit()
                break
    time.sleep(2)
    linkanitic = driver.current_url
    # Tách chuỗi URL bằng dấu "/"
    parts = linkanitic.split("/")

    # Loại bỏ phần tử cuối cùng, tức là "/analytics"
    new_link = "/".join(parts[:-1])

    print(new_link)  # In ra kết quả
    window_handles = driver.window_handles
    driver.switch_to.window(window_handles[0])
    driver.implicitly_wait(10)
    driver.find_element(By.XPATH,"//input[@placeholder='Paste link here']").send_keys(str(new_link))
    time.sleep(0.1)
    while True:
        try:
            driver.find_element(By.XPATH,"//span[text()='Verify']").click()
            break
        except:
            scroll_down_pixels = 10

            # Execute JavaScript code to scroll down
            driver.execute_script(f"window.scrollBy(0, {scroll_down_pixels});")
    time.sleep(2)
    try:
        driver.execute_script('document.querySelector("#app > div.v-dialog__content.v-dialog__content--active > div > div > div > div.footer.flex-align-center.flex-justify-end > div > div:nth-child(2) > button > span").click()')
    except:
        print("không cần")
def close_all_chrome_windows():
    try:
        # Thực hiện lệnh taskkill để đóng tất cả cửa sổ Chrome
        subprocess.run(['taskkill', '/F', '/IM', 'chrome.exe'], check=True)
        print("Đã đóng tất cả cửa sổ Chrome thành công.")
    except subprocess.CalledProcessError as e:
        print(f"Lỗi: {e}")
def login_twitterketnoi(driver, tk, mk, fa):
    # Khởi tạo trình duyệt

    

    time.sleep(1)
    driver.implicitly_wait(10)
    username922 = driver.find_element(By.NAME, "text")

    actions = ActionChains(driver)
    actions.send_keys_to_element(username922, tk).perform()
    time.sleep(0.1)
    driver.implicitly_wait(10)

    next_button = driver.find_element(By.XPATH, "//span[text()='Next']")
    actions.move_to_element(next_button)
    time.sleep(0.1)
    actions.click().perform()


    
    time.sleep(0.1)
    driver.implicitly_wait(10)
    print(mk)
    mk123 = driver.find_element(By.NAME, "password")
    time.sleep(0.1)

    actions.send_keys_to_element(mk123, mk).perform()
    time.sleep(1)
    time.sleep(0.1)
    driver.implicitly_wait(10)
    login_button = driver.find_element(By.XPATH, "//span[text()='Log in']")

    actions.move_to_element(login_button)
    time.sleep(0.1)
    actions.click().perform()
    time.sleep(1)
    driver.implicitly_wait(10)
    token2929 = driver.find_element(By.NAME,"text")
    print(fa)
    url = "https://2fa.live/tok/" + fa
    response = requests.get(url)
    data = json.loads(response.text)
    token = data["token"]

    print(token)

    token2929 = driver.find_element(By.NAME,"text")
    actions.send_keys_to_element(token2929, token).perform()
    driver.implicitly_wait(10)
    next_button = driver.find_element(By.XPATH,"//span[text()='Next']")
    actions.move_to_element(next_button)
    actions.click().perform()
    time.sleep(0.1)
    try:
        driver.find_element(By.XPATH,"//span[text()='Next']")
        driver.quit()
    except NoSuchElementException:
        print("tiep")
    time.sleep(1)
    link = driver.current_url
    if link == "https://x.com/account/access":

        tabs = driver.window_handles

        # Iterate through each tab
        for tab in tabs:
            # Switch to the tab
            driver.switch_to.window(tab)
            # Check if the current tab's URL matches the desired URL
            if driver.current_url == "https://x.com/account/access":

                # If it matches, break the loop
                break
        driver.refresh()
        time.sleep(10)
        if "Your account has been locked." in driver.page_source:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH, "//input[@type='submit' and @value='Start']").click()
            start_time = time.time()  # Record the start time

            while True:
                if "Account unlocked." in driver.page_source:
                    # If "Account unlocked." is found, click the button and break the loop
                    driver.implicitly_wait(10)
                    driver.find_element(By.XPATH, "//input[@type='submit' and @value='Continue to X']").click()
                    break
                else:
                    if "Something went wrong." in driver.page_source:
                        # If "Something went wrong." is found, navigate to Twitter's homepage
                        driver.get("https://x.com/")
                    
                    # Check if more than 60 seconds have passed
                    elapsed_time = time.time() - start_time
                    if elapsed_time > 60:
                        print("Timeout: Condition not met within 60 seconds.")
                        driver.quit()
                        break

                    # Wait for 1 second before checking again
                    time.sleep(1)
                    print("Tiếp tục xem đã giải captcha chưa")

                    # Refresh the page
                    #loccapcha(driver)
        elif "Continue to X" in driver.page_source:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH, "//input[@type='submit' and @value='Continue to X']").click()
        elif "Account unlocked" in driver.page_source:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH, "//input[@type='submit' and @value='Continue to X']").click()
        else:
            start_time = time.time()  # Record the start time
            #loccapcha(driver)
            while True:
                if "https://x.com/account/access" in driver.current_url:
                    time.sleep(10)
                    try:
                        driver.implicitly_wait(10)
                        driver.find_element(By.XPATH, "//input[@type='submit' and @value='Continue to X']").click()
                        print("Đã giải thành công")
                        break  # Break the loop if successful
                    except NoSuchElementException:
                        print("Không cần")
                else:
                    print("Điều kiện không khớp")

                # Check if more than 60 seconds have passed
                elapsed_time = time.time() - start_time
                if elapsed_time > 60:
                    print("Timeout: Condition not met within 60 seconds.")
                    driver.quit()  # Quit the driver
                    break

                time.sleep(1)  # Wait for 1 second before rechecking
    else:
        print("Tiếp tục")
    time.sleep(5)
def login_twitter(driver, tk, mk, fa):
    # Khởi tạo trình duyệt

    load_extension_page(driver,"https://x.com/")
    time.sleep(5)
    
    current_url = driver.current_url
    print(current_url)
    if "https://x.com/home" in current_url:
        print("Đã đăng nhập")
    elif current_url == "https://x.com/account/access":
        print("Tiếp")
    else:
        driver.set_page_load_timeout(600)
        load_extension_page(driver,"https://x.com/i/flow/login")
        time.sleep(1)
        driver.implicitly_wait(600)
        username922 = driver.find_element(By.NAME, "text")

        actions = ActionChains(driver)
        actions.send_keys_to_element(username922, tk).perform()
        time.sleep(0.1)
        driver.implicitly_wait(10)

        next_button = driver.find_element(By.XPATH, "//span[text()='Next']")
        actions.move_to_element(next_button)
        time.sleep(0.1)
        actions.click().perform()


        
        time.sleep(0.1)
        driver.implicitly_wait(10)
        print(mk)
        mk123 = driver.find_element(By.NAME, "password")
        time.sleep(0.1)

        actions.send_keys_to_element(mk123, mk).perform()
        time.sleep(1)
        time.sleep(0.1)
        driver.set_page_load_timeout(600)
        driver.implicitly_wait(10)
        login_button = driver.find_element(By.XPATH, "//span[text()='Log in']")

        actions.move_to_element(login_button)
        time.sleep(0.1)
        actions.click().perform()
        time.sleep(1)
        if fa is None:
            print("không cần")
        else:
            driver.implicitly_wait(600)
            token2929 = driver.find_element(By.NAME,"text")

            print(fa)
            url = "https://2fa.live/tok/" + fa
            response = requests.get(url)
            data = json.loads(response.text)
            token = data["token"]

            print(token)
            driver.set_page_load_timeout(1000)
            token2929 = driver.find_element(By.NAME,"text")
            actions.send_keys_to_element(token2929, token).perform()
            driver.implicitly_wait(10)
            next_button = driver.find_element(By.XPATH,"//span[text()='Next']")

            actions.move_to_element(next_button)
            actions.click().perform()
            time.sleep(0.1)
            try:
                driver.find_element(By.XPATH,"//span[text()='Next']")
                driver.quit()
            except NoSuchElementException:
                print("tiep")
            time.sleep(1)
    time.sleep(10)
    

    
def login_twitter1(driver, tk, mk, fa):
    # Khởi tạo trình duyệt

    
    try:
        driver.get("https://x.com/logout")
        time.sleep(1)
        driver.implicitly_wait(10)
        spans = driver.find_elements(By.TAG_NAME,'span')

        # Lặp qua các phần tử và tìm phần tử có nội dung 'Log out'
        for span in spans:
            if span.text.strip() == 'Log out':
                span.click()
                break
        time.sleep(5)
    except:
        print("on")
    time.sleep(5)
    time.sleep(0.1)
    current_url = driver.current_url
    print(current_url)
    
    num_retries = 0
    while num_retries < 3:
        num_retries += 1
        try:
            driver.get("https://x.com/login")

            time.sleep(0.1)
            driver.implicitly_wait(10)
            username922 = driver.find_element(By.NAME, "text")

            actions = ActionChains(driver)
            actions.send_keys_to_element(username922, tk).perform()
            time.sleep(0.1)
            username922.clear()
            time.sleep(0.1)
            actions.send_keys_to_element(username922, tk).perform()
            time.sleep(0.1)
            driver.implicitly_wait(10)

            next_button = driver.find_element(By.XPATH, "//span[text()='Next']")
            actions.move_to_element(next_button)
            time.sleep(0.1)
            actions.click().perform()

            break
        except Exception as e:
            print(e)
    time.sleep(0.1)
    driver.implicitly_wait(10)
    print(mk)
    mk123 = driver.find_element(By.NAME, "password")
    time.sleep(0.1)

    actions.send_keys_to_element(mk123, mk).perform()
    time.sleep(0.1)
    driver.implicitly_wait(10)
    login_button = driver.find_element(By.XPATH, "//span[text()='Log in']")

    actions.move_to_element(login_button)
    time.sleep(0.1)
    actions.click().perform()
    time.sleep(1)
    driver.implicitly_wait(10)
    token2929 = driver.find_element(By.NAME,"text")
    print(fa)
    url = "https://2fa.live/tok/" + fa
    response = requests.get(url)
    data = json.loads(response.text)
    token = data["token"]

    print(token)

    token2929 = driver.find_element(By.NAME,"text")
    actions.send_keys_to_element(token2929, token).perform()
    driver.implicitly_wait(10)
    next_button = driver.find_element(By.XPATH,"//span[text()='Next']")
    actions.move_to_element(next_button)
    actions.click().perform()
    time.sleep(5)
    try:
        driver.find_element(By.XPATH,"//span[text()='Next']")
        driver.quit()
    except:
        print("tiep")
    link = driver.current_url
    if link == "https://x.com/account/access":

        tabs = driver.window_handles

        # Iterate through each tab
        for tab in tabs:
            # Switch to the tab
            driver.switch_to.window(tab)
            # Check if the current tab's URL matches the desired URL
            if driver.current_url == "https://x.com/account/access":

                # If it matches, break the loop
                break
        driver.refresh()
        time.sleep(5)
        if "Your account has been locked." in driver.page_source:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH, "//input[@type='submit' and @value='Start']").click()
            while True:
                if "Use the arrows to find the image where the number on each ring adds up to the number on the left" in driver.page_source:
                    driver.quit()
                elif "Account unlocked." in driver.page_source:
                    driver.implicitly_wait(10)
                    driver.find_element(By.XPATH, "//input[@type='submit' and @value='Continue to X']").click()
                    break
                else:
                    time.sleep(5)
                    print("Tiếp tục xem đã giải capcha chưa")
        elif "Continue to X" in driver.page_source:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH, "//input[@type='submit' and @value='Continue to X']").click()
        elif "Account unlocked" in driver.page_source:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH, "//input[@type='submit' and @value='Continue to X']").click()
        else:
            #
            while True:
                if "https://x.com/account/access" in driver.current_url:
                    time.sleep(5)
                    try:
                        driver.implicitly_wait(10)
                        driver.find_element(By.XPATH, "//input[@type='submit' and @value='Continue to X']").click()
                    except NoSuchElementException:
                        print("Không cần")
                else:
                    print("Đã giải thành công")
                    break

            print("Điều kiện không khớp")
            time.sleep(0.1)
        time.sleep(0.1)
    else:
        print("Tiếp tục")
    time.sleep(0.1)
def kiemtracapcha(driver):
    link = driver.current_url
    if link == "https://x.com/account/access":

        tabs = driver.window_handles

        # Iterate through each tab
        for tab in tabs:
            # Switch to the tab
            driver.switch_to.window(tab)
            # Check if the current tab's URL matches the desired URL
            if driver.current_url == "https://x.com/account/access":

                # If it matches, break the loop
                break
        driver.refresh()
        time.sleep(5)
        if "Your account has been locked." in driver.page_source:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH, "//input[@type='submit' and @value='Start']").click()
            while True:
                if "Use the arrows to find the image where the number on each ring adds up to the number on the left" in driver.page_source:
                    driver.quit()
                elif "Account unlocked." in driver.page_source:
                    driver.implicitly_wait(10)
                    driver.find_element(By.XPATH, "//input[@type='submit' and @value='Continue to X']").click()
                    break
                else:
                    time.sleep(5)
                    print("Tiếp tục xem đã giải capcha chưa")
        elif "Continue to X" in driver.page_source:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH, "//input[@type='submit' and @value='Continue to X']").click()
        elif "Account unlocked" in driver.page_source:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH, "//input[@type='submit' and @value='Continue to X']").click()
        else:
            #
            while True:
                if "https://x.com/account/access" in driver.current_url:
                    time.sleep(5)
                    try:
                        driver.implicitly_wait(10)
                        driver.find_element(By.XPATH, "//input[@type='submit' and @value='Continue to X']").click()
                    except NoSuchElementException:
                        print("Không cần")
                else:
                    print("Đã giải thành công")
                    break

            print("Điều kiện không khớp")
            time.sleep(0.1)
        time.sleep(0.1)
    else:
        print("Tiếp tục")
def kiemtra(driver,tk,mk,fa,noidung):
    while True:
        try:
            driver.find_element(By.XPATH,f"//span[text()='{noidung}']")
            break
        except:
            print("đợi")
            driver.refresh()
            time.sleep(0.1)
def click(driver,noidung):
    while True:
        try:
            driver.find_element(By.XPATH,f"//*[contains(text(), '{noidung}')]").click()
            break
        except:
            scroll_down_pixels = 10

            # Execute JavaScript code to scroll down
            driver.execute_script(f"window.scrollBy(0, {scroll_down_pixels});")
def connectsau(driver, tk, mk, fa):
    if "Connect your account" in driver.page_source:
        try:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,"//div//span[text()='Connect X']").click()
            time.sleep(5)
            all_handles = driver.window_handles
            for handle in all_handles:
                driver.switch_to.window(handle)
                
                # Check if the desired element exists in the current tab
                if "https://api.twitter.com/oauth/authorize?oauth_token=" in driver.current_url:
                    linkt = driver.current_url
                    print("ok")
                    time.sleep(0.1)
                    try:
                        driver.implicitly_wait(10)
                        driver.find_element(By.NAME,"session[username_or_email]").send_keys(str(tk))
                        time.sleep(0.1)
                        login_twitter1(driver, tk, mk, fa)
                        driver.get(str(linkt))
                        #time.sleep(10)
                    except NoSuchElementException:
                        print("kong can")
                    time.sleep(0.1)
                    while True:
                        try:
                            driver.execute_script('document.querySelector("#allow").click()')
                            break
                        except:
                            print("ok")
                    time.sleep(5)
                    try:
                        if "This account is suspended." in driver.page_source:
                            driver.quit()
                        else:
                            print("tiếp")
                    except:
                        print("xong")
            all_handles = driver.window_handles
            driver.switch_to.window(all_handles[0])
            time.sleep(30)
            driver.find_element(By.XPATH,"//span[text()='Verify']").click()
            # time.sleep(10)
            # while True:
            #     try:
            #         # Kiểm tra xem phần tử "Verify" có tồn tại hay không
            #         element = driver.find_element(By.XPATH, "//span[text()='Verify']")
            #         time.sleep(5)
            #         element.click()
                    
                    
            #     except:
            #         print("đã very")
            #         break
            #         # Nếu không tìm thấy, tăng biến đếm lên 1
                    
            
        except NoSuchElementException:
            print("khoong caan")
    else:
        
        print("da connect")
def folow(driver, tk, mk, fa):
    try:
        
        all_handles = driver.window_handles

        # Iterate through each tab
        for handle in all_handles:
            # Switch to the tab
            driver.switch_to.window(handle)

            # Check if the current tab's URL matches the expected one
            if "follow" in driver.current_url:
                time.sleep(2)
                print("ok")
                link1 = driver.current_url
                try:
                    driver.implicitly_wait(10)
                    folow = driver.find_element(By.XPATH,"//span[text()='Follow']")
                except:
                    login_twitter1(driver, tk, mk, fa)
                    time.sleep(0.1)
                    driver.get(str(link1))
                if "Want to log in first?" in driver.page_source:
                    
                    login_twitter1(driver, tk, mk, fa)
                    time.sleep(0.1)
                    driver.get(str(link1))
                #time.sleep(5)
                
                driver.implicitly_wait(10)
                folow = driver.find_element(By.XPATH, "//span[text()='Follow']")
                time.sleep(2)

                # Sử dụng execute_script để click
                driver.execute_script("arguments[0].click();", folow)

                                
                break  # Exit the loop if the condition is met
        
        
        time.sleep(5)
        
        
        

        
    except NoSuchElementException:
        print("đã folow thành công sẵn rồi")
def ketnoitruoc(driver,tk,mk,fa,noidung):
    try:
        driver.implicitly_wait(1)
        driver.find_element(By.XPATH,"//div//span[text()='Connect X']").click()
        time.sleep(0.1)
        all_handles = driver.window_handles
        for handle in all_handles:
            driver.switch_to.window(handle)
            
            # Check if the desired element exists in the current tab
            if "https://api.twitter.com/oauth/authorize?oauth_token=" in driver.current_url:
                linkt = driver.current_url
                print("ok")
                time.sleep(0.1)
                try:
                    driver.implicitly_wait(10)
                    driver.find_element(By.NAME,"session[username_or_email]").send_keys(str(tk))
                    time.sleep(0.1)
                    login_twitter1(driver, tk, mk, fa)
                    driver.get(str(linkt))
                    time.sleep(2)
                except NoSuchElementException:
                    print("kong can")
                driver.execute_script('document.querySelector("#allow").click()')
                time.sleep(5)
                try:
                    if "This account is suspended." in driver.page_source:
                        driver.quit()
                    else:
                        print("tiếp")
                except:
                    print("xong")
                window_handles = driver.window_handles
                driver.switch_to.window(window_handles[0])
                time.sleep(0.1)
                driver.implicitly_wait(10)
                driver.find_element(By.XPATH,f"//span[text()='{noidung}']").click()

    except NoSuchElementException:
        print("khoong caan")
def like(driver, tk, mk, fa,linknoidung):
    
        
        
    while True:
        try:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,"//span[text()='Like']").click()
            break
        except:
            time.sleep("lỗi like")
            driver.refresh()

       

def repost(driver):

        
        
    
    while True:
        try:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,"//span[text()='Repost']").click()
            break
        except:
            print("lỗi reposst")
            driver.refresh()

    
    
    
    time.sleep(0.1)

def sendvideo(driver):
    driver.get("https://studio.taffystream.io/video-library/saved-clips")
    # while True:
    #     if driver.current_url == "https://studio.taffystream.io/video-library/saved-clips":
    #         break
    #     else:
    #         print("ok")
    driver.implicitly_wait(500)
    input_file = driver.find_element(By.CSS_SELECTOR,'.video-library-tabs input[type="file"]')
        # Đường dẫn tới thư mục chứa video
    video_directory = f'E:\KEO GLX GOLOGIN\taivideoytb'

    # # Lọc tất cả các tệp có định dạng mp4 trong thư mục
    # mp4_files = [file for file in os.listdir(video_directory) if file.endswith('.mp4')]

    # # Chọn ngẫu nhiên một tệp từ danh sách các tệp mp4
    # random_mp4_file = random.choice(mp4_files)

    # # Tên tệp được chọn ngẫu nhiên
    # print("Tên của tệp MP4 được chọn ngẫu nhiên:", random_mp4_file)
    while True:
        try:
            relative_video_path = f"{random.randint(1,200)}.mp4"

            # Chuyển đổi đường dẫn tương đối thành đường dẫn tuyệt đối
            absolute_video_path = os.path.abspath(relative_video_path)
            # Gửi đường dẫn video vào input
            input_file.send_keys(absolute_video_path)
            driver.implicitly_wait(500)
            driver.find_element(By.XPATH,"//button[contains(., 'Mint')]").click()
            time.sleep(2)
            try:
                driver.execute_script('return document.querySelector("body > w3m-modal").shadowRoot.querySelector("wui-flex > wui-card > w3m-router").shadowRoot.querySelector("div > w3m-networks-view").shadowRoot.querySelector("wui-grid > wui-card-select").shadowRoot.querySelector("button > wui-network-image").click()')
                time.sleep(2)
                driver.execute_script('document.querySelector("body > w3m-modal").shadowRoot.querySelector("wui-flex > wui-card > w3m-router").shadowRoot.querySelector("div > w3m-connect-view").shadowRoot.querySelector("wui-flex > wui-list-wallet:nth-child(2)").shadowRoot.querySelector("button > wui-text").click()')
                
                time.sleep(5)
            except:
                print("ok")
            
            while True:
                try:
                    window_handles = driver.window_handles
                    driver.switch_to.window(window_handles[-1])
                    try:
                        driver.implicitly_wait(2)
                        driver.find_element(By.XPATH,"//button[text()='Got it']").click()
                    except:
                        print("không cần")
                    time.sleep(1)
                    script = "document.querySelector('.button.btn--rounded.btn-primary').click();"

                    # Execute the script
                    driver.execute_script(script)
                    time.sleep(1)

                except:
                    print("đã kết nối thành công")
                    break
            time.sleep(0.1)
            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[0])
            try:
                driver.implicitly_wait(500)
                driver.find_element(By.XPATH,"//button[contains(., 'Mint')]").click()
            except:
                print("pkktkgkg")
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,"//button[contains(text(), 'Next Step')]").click()
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,"//input[@type='checkbox']").click()
            driver.implicitly_wait(500)
            driver.find_element(By.XPATH,"//button[contains(., 'Mint')]").click()
            break
        except:
            print("ko")
def dangkitwitersmailpro(driver,link332de2):
    # load_extension_page(driver,"chrome://extensions/?id=pdapfegipbhcbdkalhidkknidpcjicbk")
    # time.sleep(2)
    # driver.execute_script(
    #         "chrome.management.setEnabled('pdapfegipbhcbdkalhidkknidpcjicbk', false, function() {"
    #         "    if (chrome.runtime.lastError) {"
    #         "        console.error(chrome.runtime.lastError.message);"
    #         "    } else {"
    #         "        console.log('Extension disabled successfully');"
    #         "    }"
    #         "});"
    #     )
    while True:
        try:
            with open('proxyxoay.txt', 'r') as file:
                lines = file.readlines()
            # Choose a random line from the list of lines
            random_line = random.choice(lines)

            # Print the randomly chosen line
            print(random_line)
            proxies = {
                'https': 'http://' + str(random_line.strip())
            }
            # with open('link.txt', 'r') as file:
            #     url_from_file = file.read().strip()

            # driver.get(str(url_from_file))
            # driver.implicitly_wait(10)
            # driver.find_element(By.ID,"api-key").send_keys("e72fbf58a85b4d63b2e5e4d71ca0c737")
            # driver.implicitly_wait(10)
            # driver.find_element(By.ID,"save-btn").click()
            # time.sleep(2)
            while True:
                try:
                    ten = ["Tran", "Nguyen", "Le", "Pham", "Huynh", "Hoang", "Phan", "Vu", "Dang", "Bui", "Do", "Ho", "Ngo", "Duong", "Ly", "Dao", "Dinh", "Doan", "Vuong", "Trinh"]
                    tenmail = random.choice(ten)
                    load_extension_page(driver,"https://x.com/i/flow/signup")
                    
                    load_extension_page(driver,"https://x.com/i/flow/signup")
                    #driver.refresh()
                    driver.implicitly_wait(30)
                    button = driver.find_element(By.XPATH,"//button[@role='button']//span[contains(text(), 'Create account')]")
                    time.sleep(0.1)
                    button.click()
                    
                    names = ['tai vo','khanh nhi','minh tuyet','thu thao','thanh tuyet','minh tri','van tai','thu ha','tu nguyen','huy hoang']
                    random_name = random.choice(names)
                    random_name_no_spaces = random_name.replace(" ", "")
                    print(random_name_no_spaces)
                    driver.implicitly_wait(10)
                    driver.find_element(By.NAME,"name")
                    driver.refresh()
                    driver.implicitly_wait(30)
                    button = driver.find_element(By.XPATH,"//button[@role='button']//span[contains(text(), 'Create account')]")
                    time.sleep(0.1)
                    button.click()
                    driver.implicitly_wait(10)
                    driver.find_element(By.NAME,"name").send_keys(str(random_name))
                    
                    
                    # URL và headers cho yêu cầu
                    url = "https://temporary-email3.p.rapidapi.com/get"
                    headers = {
                        "X-RapidAPI-Key": "85c0287154msh55bbcae0f065001p12e225jsn7770b94568bb",
                        "X-RapidAPI-Host": "temporary-email3.p.rapidapi.com"
                    }

                    # Tham số cho yêu cầu
                    params = {
                        "username": "random",
                        "domain": "random.com"
                    }

                    # Gửi yêu cầu GET
                    response = requests.get(url, headers=headers, params=params)

                    # Kiểm tra phản hồi và lấy email
                    if response.status_code == 200:
                        data = response.json()  # Chuyển đổi phản hồi sang JSON
                        email = data.get('items', {}).get('email')  # Lấy email từ phản hồi
                        if email:
                            print("Email:", email)
                        else:
                            print("Email not found in the response.")
                    else:
                        print(f"Failed to fetch data. Status Code: {response.status_code}")
                        print("Response Content:", response.text)
                    email_address = email
                    driver.implicitly_wait(10)
                    #driver.find_element(By.XPATH,"//button[@role='button']//span[contains(text(), 'Use email instead')]").click()

                    driver.find_element(By.NAME,"email").send_keys(email_address)
                    time.sleep(0.1)
                    select_element = driver.find_element(By.ID, 'SELECTOR_1')

                    # Create a Select object
                    select = Select(select_element)

                    # Define the list of month values, excluding February (value="2")
                    months = [str(i) for i in range(1, 13) if i != 2]

                    # Choose a random month from the list
                    random_month = random.choice(months)

                    # Select the random month
                    select.select_by_value(random_month)
                    time.sleep(0.1)
                    select_element = driver.find_element(By.ID, 'SELECTOR_2')

                    # Create a Select object
                    select = Select(select_element)

                    # Define the list of day values (1 to 31)
                    days = [str(i) for i in range(1, 32)]

                    # Choose a random day from the list
                    random_day = random.choice(days)

                    # Select the random day
                    select.select_by_value(random_day)

                    # Optionally, print the selected day
                    print(f'Selected day: {random_day}')
                    time.sleep(0.1)
                    select_element = driver.find_element(By.ID, 'SELECTOR_3')

                    # Create a Select object
                    select = Select(select_element)

                    # Define the list of year values (from 2000 to 2003)
                    years = ['1980', '1981', '1982', '1983','1984','1985','1986','1987','1988','1989']

                    # Choose a random year from the list
                    random_year = random.choice(years)

                    # Select the random year
                    select.select_by_value(random_year)

                    # Optionally, print the selected year
                    print(f'Selected year: {random_year}')
                    time.sleep(5)


                    driver.implicitly_wait(10)
                    button = driver.find_element(By.XPATH, "//button[@role='button' and contains(@data-testid, 'ocfSignupNextLink')]//span[contains(text(), 'Next')]")
                    driver.execute_script("arguments[0].click();", button)

                    
                    break
                except NoSuchElementException:
                    time.sleep(1)

            start_time = time.time()
            timeout = 30  # seconds
            time.sleep(5)
           

            while True:
                if "We sent you a code" in driver.page_source:
                    
                    break
                elif "Happening now" in driver.page_source:
                    driver.quit()
                    break
                if time.time() - start_time > 60:
                    # Tải lại iframe nếu đã vượt quá 60 giây và điều kiện chưa thỏa mãn
                    driver.execute_script("""
                        // Tìm iframe bằng ID
                        var iframe = document.getElementById('arkoseFrame');
                        // Tải lại iframe bằng cách đặt lại thuộc tính src
                        iframe.src = iframe.src;
                    """)
                    # Cập nhật lại thời gian bắt đầu để tiếp tục đợi thêm 60 giây
                    start_time = time.time()
                

            start_time = time.time()
            timeout = 30  # seconds
            while True:

                url = "https://temporary-email3.p.rapidapi.com/check"
                headers = {
                    "X-RapidAPI-Key": "85c0287154msh55bbcae0f065001p12e225jsn7770b94568bb",
                    "X-RapidAPI-Host": "temporary-email3.p.rapidapi.com"
                }

                # Tham số cho yêu cầu
                params = {
                    "email": email,
                    "timestamp": "1729049736"
                }

                # Gửi yêu cầu GET
                response = requests.get(url, headers=headers, params=params)

                # Kiểm tra phản hồi và xử lý kết quả
                if response.status_code == 200:
                    data = response.json()  # Chuyển đổi phản hồi sang JSON
                    print("Response JSON:", data)  # In ra toàn bộ phản hồi
                    
                    # Lấy giá trị textSubject từ items
                    items = data.get('items', [])
                    if items:
                        text_subject = items[0].get('textSubject', '')
                        
                        # Sử dụng regex để lấy 6 chữ số từ textSubject
                        match = re.search(r'\b(\d{6})\b', text_subject)
                        if match:
                            code = match.group(1)
                            print("6-digit code:", code)
                            break
                        else:
                            print("No 6-digit code found in the subject.")
                            
                    else:
                        if time.time() - start_time > timeout:
                            print("Timeout exceeded")
                            
                            driver.find_element(By.CSS_SELECTOR,'button[aria-label="Back"]').click()
                            
                            while True:
                                if "We sent you a code" in driver.page_source:
                                    start_time = time.time()
                                    break
                        else:
                            print("doi")
                else:
                    print(f"Failed to fetch data. Status Code: {response.status_code}")
                    print("Response Content:", response.text)
            driver.find_element(By.NAME,"verfication_code").send_keys(code)
            time.sleep(2)
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,"//button[@role='button']//span[contains(text(), 'Next')]").click()
            driver.implicitly_wait(30)
            pass123 = random_password = ''.join(random.choice(string.ascii_letters + string.digits) for _ in range(10))
            driver.find_element(By.NAME,"password").send_keys(str(pass123))
            time.sleep(2)
            driver.implicitly_wait(10)
            element = driver.find_element(
                By.XPATH, "//button[@role='button' and contains(@data-testid, 'LoginForm_Login_Button')]//span[contains(text(), 'Sign up')]"
            )

            # Sử dụng ActionChains để nhấp vào phần tử
            actions = ActionChains(driver)
            actions.move_to_element(element).click().perform()
            time.sleep(2)
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,"//button[@role='button' and contains(@data-testid, 'ocfSelectAvatarSkipForNowButton')]//span[contains(text(), 'Skip for now')]").click()

                
            # print(numbers[0])

                
            
            
            # driver.implicitly_wait(60)
            # driver.find_element(By.XPATH,"//button[@role='button' and contains(@data-testid, 'ocfSelectAvatarSkipForNowButton')]//span[contains(text(), 'Skip for now')]").click()
            # driver.implicitly_wait(60)
            # driver.find_element(By.XPATH,"//button[@role='button' and contains(@data-testid, 'ocfSelectAvatarSkipForNowButton')]//span[contains(text(), 'Skip for now')]").click()


            load_extension_page(driver,"https://x.com/")
            time.sleep(2)
            if "Happening now" in driver.page_source:
                #happend
                print("tiếp tục")
                #driver.quit()
                
            else:
                start_time = time.time()
                while True:
                    try:
                        driver.refresh()
                        driver.implicitly_wait(10)
                        
                        # Click on the profile link
                        driver.find_element(By.XPATH, '//a[@data-testid="AppTabBar_Profile_Link"]').click()
                        time.sleep(5)
                        
                        # Extract the current URL and parse the username
                        current_url = driver.current_url
                        url_parts = current_url.split('/')
                        usename = url_parts[3]
                        
                        # Define the filename
                        filename = "twiternew.txt"
                        
                        # Write or append the content to the file
                        content = f"{usename}|{pass123}|{email_address}|thongvo123|{link332de2}"
                        
                        if not os.path.isfile(filename):
                            with open(filename, "w", encoding="utf8") as f:
                                f.write(content)
                        else:
                            with open(filename, "a", encoding="utf8") as f:
                                f.write("\n" + content)
                        
                        break  # Exit the loop if successful
                    except NoSuchElementException:
                        print("Element not found. Retrying...")
                    except Exception as e:
                        print(f"An error occurred: {e}")
                    
                    # Check if the elapsed time is more than 30 seconds
                    if time.time() - start_time > 30:
                        print("Operation timed out after 30 seconds.")
                        driver.quit()
                        break
                break
        except NoSuchElementException:
            print("thử lại sao 5s")
            time.sleep(5)
            

    return usename,pass123
def dangki(driver,mail123,diachivi):
    #driver = webdriver.Chrome()
    driver.get("https://studio.taffystream.io/signup")
    time.sleep(5)
    if driver.current_url == "https://studio.taffystream.io/home":
        print("da mint roi")
        sendvideo(driver)
        time.sleep(60)
        guiform(driver,mail123,diachivi)
    else:
        ten = ["Tran", "Nguyen", "Le", "Pham", "Huynh", "Hoang", "Phan", "Vu", "Dang", "Bui", "Do", "Ho", "Ngo", "Duong", "Ly", "Dao", "Dinh", "Doan", "Vuong", "Trinh"]
        ho = ["Van", "Thi", "Huu", "Tu", "Anh", "Thanh", "Duc", "Ngoc", "Cong", "Quoc", "Nhu", "Gia", "Minh", "Hoai", "Hai", "Hanh", "Hoa", "Phuoc", "Tien", "Tam"]

        # Chọn một tên và một họ ngẫu nhiên
        ten_random = random.choice(ten)
        ho_random = random.choice(ho)

        print("Tên ngẫu nhiên:", ten_random)
        print("Họ ngẫu nhiên:", ho_random)
        driver.implicitly_wait(10)
        driver.find_element(By.NAME,"first_name").send_keys(str(ten_random))
        driver.implicitly_wait(10)
        driver.find_element(By.NAME,"last_name").send_keys(str(ho_random))
        driver.implicitly_wait(10)
        driver.find_element(By.NAME,"username").send_keys(str(ten_random) + str(random.randint(111,99999)))
        driver.implicitly_wait(10)
        driver.find_element(By.NAME,"email").send_keys(str(mail123))
        # Định nghĩa tập hợp các ký tự có thể sử dụng
        length = 12

        # Tạo một tập hợp chứa tất cả các ký tự có thể sử dụng
        characters = string.ascii_letters + string.digits + string.punctuation

        # Tạo một mật khẩu ngẫu nhiên bằng cách lựa chọn ngẫu nhiên từ tập hợp các ký tự
        strong_password = ''.join(random.choice(characters) for _ in range(length))

        print(strong_password)

        print(strong_password)
        upper_case_chars = string.ascii_uppercase

        # Chọn một ký tự viết hoa ngẫu nhiên từ chuỗi trên
        random_upper_case_char = random.choice(upper_case_chars)
        matkhau = "@" + str(random_upper_case_char) + str(strong_password) + str(random.randint(1,1000))
        driver.implicitly_wait(10)
        driver.find_element(By.NAME,"password").send_keys(str(matkhau))
        driver.implicitly_wait(10)
        driver.find_element(By.NAME,"confirm-password").send_keys(str(matkhau))
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//button[contains(text(), 'Next')]").click()
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//h5[text()='Free']").click()
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,'//input[@aria-label="Accept terms and conditions"]').click()
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,'//button[text()="Agree"]').click()
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//button[contains(text(), 'Next')]").click()
        time.sleep(10)
        import requests
        import json
        url = "https://api.mail.tm/token"

        payload = {
            "address":  str(mail123),
            "password": "thongvo123"
        }

        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json"
        }

        response = requests.post(url, headers=headers, json=payload)
        print(response.json())
        thu = response.json()['token']
        print(thu)
        header = {"authorization": "Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiJ9.eyJpYXQiOjE2Njk5Njc3ODMsInJvbGVzIjpbIlJPTEVfVVNFUiJdLCJ1c2VybmFtZSI6Im5wbWtjbGlqZEBrYXJlbmtleS5jb20iLCJpZCI6IjYzODlhZmE3YTY5ZDBmZmMyMTE0MzIxYyIsIm1lcmN1cmUiOnsic3Vic2NyaWJlIjpbIi9hY2NvdW50cy82Mzg5YWZhN2E2OWQwZmZjMjExNDMyMWMiXX19.-odZWTOPkc1txvfqWXG3GoBAyWwI9q_copAGufZFcC1qTWR4gEiuDwJiZEwgOfpS0uBaaiWrsCTC4ubS9a3_5g"}
        header["authorization"] = "Bearer " + thu
        print(header)
        r = requests.get('https://api.mail.tm/messages', headers=header)
        mail_katusu = r.text
        mail_katusu = json.loads(mail_katusu)
        if str(mail_katusu["hydra:member"]) != "[]":
            id = mail_katusu["hydra:member"][0]["id"]
            r = requests.get(f'https://api.mail.tm/messages/{id}', headers=header)
            mail = r.text
            mail = json.loads(mail)
            print(mail)
            text = mail["html"]
            for text1 in text:
                so = re.findall(r'\d', text1)
                so_sau_khi_loc = ''.join(so[:6])
                print("Số sau khi lọc:", so_sau_khi_loc)
        i = 1

        for i in range(6):
            driver.implicitly_wait(10)
            driver.find_element(By.ID,f"confirmCodeInput{i}").send_keys(str(so_sau_khi_loc[i]))
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//button[contains(text(), 'Next')]").click()
        time.sleep(5)
        
        sendvideo(driver)
        time.sleep(60)
        guiform(driver,mail123,diachivi)
        
    # driver.get("https://studio.taffystream.io/video-library/saved-clips")
    # video_path = "taivo.mp4"

    # # Xác định phần tử input file
    # input_file = driver.find_element(By.CSS_SELECTOR,'input[type="file"]')
    # #driver.find_element(By.CSS_SELECTOR,"")
    # # Gửi đường dẫn của video vào phần tử input file
    # input_file.send_keys(video_path)

    # # Chờ một khoảng thời gian để video được tải lên (thời gian có thể thay đổi tùy thuộc vào tốc độ mạng)
    # time.sleep(100000)
def guiform(driver,mail123,diachivi):
    driver.get("https://forms.gle/f13Ycs23G25K3iGS8")
    driver.implicitly_wait(10)
    input_elements = driver.find_elements(By.CSS_SELECTOR,'input.whsOnd.zHQkBf')
    input_elements[0].send_keys(str(diachivi))
    time.sleep(1)
    input_elements[1].send_keys(str(mail123))
    driver.implicitly_wait(10)
    driver.find_element(By.XPATH,'//span[text()="Gửi"]').click()
    time.sleep(10)
    
def nvtuychon(driver, tk, mk, fa, linknoidung):
    try:
        tabs = driver.window_handles
        tab_matched = False  # Biến đánh dấu nếu có tab nào khớp với liên kết

        # Lặp qua từng tab và chuyển đến
        for tab in tabs:
            # Chuyển đến tab hiện tại
            driver.switch_to.window(tab)
            
            # Kiểm tra URL của tab hiện tại
            if "https://app.galxe.com/link?redirect" in driver.current_url:
                print("Tab đã được chuyển đến với URL mong muốn:", driver.current_url)
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(0.1)
                driver.implicitly_wait(10)
                driver.find_element(By.XPATH,'//*[@id="app"]/div/main/div/div/div/div/div[1]/p/button').click()
                time.sleep(2)
                all_handles = driver.window_handles
                driver.switch_to.window(all_handles[0])
                tab_matched = True  # Đặt biến đánh dấu thành True
                break  # Thoát khỏi vòng lặp sau khi tìm thấy tab khớp

        if not tab_matched:
            print("Không có tab nào khớp với liên kết:", linknoidung)

    except Exception as e:
        print("Có lỗi xảy ra:", e)

def post(driver,tk,mk,fa,linknoidung):

    all_handles = driver.window_handles

    # Iterate through each tab
    for handle in all_handles:
        # Switch to the tab
        driver.switch_to.window(handle)

        # Check if the current tab's URL matches the expected one
        if "post" in driver.current_url:
            print("ok")
            time.sleep(2)
            link1 = driver.current_url
            try:
                driver.implicitly_wait(60)
                folow = driver.find_element(By.XPATH,"//span[text()='Post']")
            except:
                login_twitter1(driver, tk, mk, fa)
                time.sleep(0.1)
                driver.get(str(link1))
            if "Want to log in first?" in driver.page_source:
                
                login_twitter1(driver, tk, mk, fa)
                time.sleep(0.1)
                driver.get(str(link1))
            #time.sleep(5)
            # driver.implicitly_wait(10)
            # driver.find_element(By.XPATH,"//div[@aria-label='Post text']").send_keys(" @FacBrosG")

                
            
            js_script = """
            // Sử dụng XPath để chọn phần tử span chứa văn bản "Post"
            let xpath = "//span[text()='Post']";
            let element = document.evaluate(xpath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;

            if (element) {
            element.click();
            } else {
            console.log("Không tìm thấy phần tử span với văn bản 'Post'.");
            }
            """

            # Thực thi mã JavaScript
            driver.execute_script(js_script)
            time.sleep(5)
               
            break  # Exit the loop if the condition is met
    
    time.sleep(0.1)
def post1(driver):

    while True:
        try:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,"//div[@aria-label='Post text']").send_keys(" @FacBrosG")
            time.sleep(5)
            js_script = """
            // Sử dụng XPath để chọn phần tử span chứa văn bản "Post"
            let xpath = "//span[text()='Post']";
            let element = document.evaluate(xpath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
            element.click();
            """

            # Thực thi mã JavaScript
            driver.execute_script(js_script)
            break
        except:
            driver.refresh()
    time.sleep(5)
def postbai(driver,tk,mk,fa,linknoidung):
    try:
        all_handles = driver.window_handles

        # Iterate through each tab
        for handle in all_handles:
            # Switch to the tab
            driver.switch_to.window(handle)

            # Check if the current tab's URL matches the expected one
            if "post" in driver.current_url:
                print("ok")
                time.sleep(2)
                link1 = driver.current_url
                driver.implicitly_wait(10)
                if "Want to log in first?" in driver.page_source:
                    
                    login_twitter1(driver, tk, mk, fa)
                    time.sleep(0.1)
                    driver.get(str(link1))
                #time.sleep(5)
                driver.implicitly_wait(10)
                driver.find_element(By.XPATH,'//*[@aria-label="Post text"]').send_keys("@FLOTUS  @aiFilmAcademy ")
                time.sleep(1)
                driver.implicitly_wait(10)
                driver.find_element(By.XPATH,"//span[text()='Post']").click()
                break  # Exit the loop if the condition is met
        
        time.sleep(0.1)
        
    except:
        print("đã post lần 1 rồi")
def close_all_tabs_except_main(driver):
    # Lấy cửa sổ hiện tại
    current_window = driver.current_window_handle
    
    # Lấy danh sách tất cả các cửa sổ
    all_windows = driver.window_handles
    
    # Đóng tất cả các cửa sổ trừ cửa sổ hiện tại
    for window in all_windows:
        if window != current_window:
            driver.switch_to.window(window)
            driver.close()
    
    # Chuyển về cửa sổ chính
    driver.switch_to.window(current_window)
def xacminhhoanthanh(driver):
    script = """
    // Tìm tất cả các phần tử <span> thỏa mãn điều kiện
    const xpath = "//span[contains(@class, '[&_svg]:text-inherit') and contains(@class, '[&_svg_path]:fill-current') and contains(@class, '[&_svg]:h-[1em]') and contains(@class, 'h-[1em]') and contains(@class, '[&_svg]:w-[1em]') and contains(@class, 'w-[1em]') and contains(@class, 'text-size-20') and contains(@class, 'hover:text-text-linkBase')]";

    // Thực hiện tìm kiếm bằng XPath
    const spans = document.evaluate(xpath, document, null, XPathResult.ORDERED_NODE_SNAPSHOT_TYPE, null);

    // Lặp qua mỗi phần tử <span> và click vào nó
    for (let i = 0; i < spans.snapshotLength; i++) {
        spans.snapshotItem(i).click();
    }
    """

    # Thực hiện JavaScript code
    driver.execute_script(script)


            


def ki(driver):
    while True:
        try:
            window_handles = driver.window_handles
            driver.switch_to.window(window_handles[-1])
            try:
                driver.implicitly_wait(2)
                driver.find_element(By.XPATH,"//button[text()='Got it']").click()
            except:
                print("không cần")
            time.sleep(1)
            script = "document.querySelector('.button.btn--rounded.btn-primary').click();"

            # Execute the script
            driver.execute_script(script)
            time.sleep(1)

        except:
            print("đã kết nối thành công")
            break
    time.sleep(0.1)
    window_handles = driver.window_handles
    driver.switch_to.window(window_handles[0])
def themail(driver):
    with open('proxyxoay.txt', 'r') as file:
        lines = file.readlines()
    # Choose a random line from the list of lines
    random_line = random.choice(lines)

    # Print the randomly chosen line
    print(random_line)
    proxies = {
        'https': 'http://192.168.1.11:' + str(random_line.strip())
    }
    ten = ["Tran", "Nguyen", "Le", "Pham", "Huynh", "Hoang", "Phan", "Vu", "Dang", "Bui", "Do", "Ho", "Ngo", "Duong", "Ly", "Dao", "Dinh", "Doan", "Vuong", "Trinh"]
    tenmail = random.choice(ten)
    tenmien = "@belgianairways.com"
    
    url = "https://api.mail.tm/accounts"
    ten1 = str(tenmail) + str(random.randint(111,999999))
    ten2 = ''.join(random.choices(string.ascii_lowercase + string.digits, k=10))
    email_address = ten2 + tenmien
    payload = {
        "address": email_address,
        "password": "thongvo123"
    }
    headers = {
        "accept": "application/ld+json",
        "Content-Type": "application/ld+json"
    }

    response = requests.post(url, json=payload, headers=headers,proxies=proxies)

    print(response.json())
    url = "https://api.mail.tm/token"
    payload = {
        "address": email_address,
        "password": "thongvo123"
    }

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    response = requests.post(url, headers=headers, json=payload,proxies=proxies)
    print(response.json())
    




    url = "https://api.mail.tm/token"

    payload = {
        "address":  email_address,
        "password": "thongvo123"
    }

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    response = requests.post(url, headers=headers, json=payload,proxies=proxies)
    print(response.json())
    while True:
        try:
            driver.implicitly_wait(10)
            driver.find_element(By.CSS_SELECTOR,"input[placeholder='Email address']").send_keys(email_address)
            break
        except:
            driver.refresh()
            driver.implicitly_wait(10)
            mai339292 = driver.find_element(By.XPATH,"//a[@href='/accountSetting/social' and contains(text(),'Social Account')]")
            time.sleep(1)
            mai339292.click()
    time.sleep(0.1)
    while True:
        try:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,"//button[text()='Send a code']").click()
            break
        except:
            print("lỗi send mail")
    time.sleep(10)

    while True:

        thu = response.json()['token']
        print(thu)
        header = {"authorization": "Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiJ9.eyJpYXQiOjE2Njk5Njc3ODMsInJvbGVzIjpbIlJPTEVfVVNFUiJdLCJ1c2VybmFtZSI6Im5wbWtjbGlqZEBrYXJlbmtleS5jb20iLCJpZCI6IjYzODlhZmE3YTY5ZDBmZmMyMTE0MzIxYyIsIm1lcmN1cmUiOnsic3Vic2NyaWJlIjpbIi9hY2NvdW50cy82Mzg5YWZhN2E2OWQwZmZjMjExNDMyMWMiXX19.-odZWTOPkc1txvfqWXG3GoBAyWwI9q_copAGufZFcC1qTWR4gEiuDwJiZEwgOfpS0uBaaiWrsCTC4ubS9a3_5g"}
        header["authorization"] = "Bearer " + thu
        print(header)
        r = requests.get('https://api.mail.tm/messages', headers=header,proxies=proxies)
        mail_katusu = r.text
        mail_katusu = json.loads(mail_katusu)
        if str(mail_katusu["hydra:member"]) != "[]":
            id = mail_katusu["hydra:member"][0]["id"]
            r = requests.get(f'https://api.mail.tm/messages/{id}', headers=header,proxies=proxies)
            mail = r.text
            mail = json.loads(mail)
            text = mail["text"]
            print(text)
            numbers = re.findall(r'\d{6}', text)
            break
        else:
            time.sleep(5)
            print("thực hiện lại")
        



    driver.implicitly_wait(10)
    driver.find_element(By.CSS_SELECTOR, "input[placeholder='Enter code']").send_keys(numbers[0])
    
    time.sleep(2)
    driver.implicitly_wait(10)
    driver.find_element(By.XPATH,"//button[.//div[text()='Verify']]").click()
    time.sleep(10)
    
    driver.refresh()

    

def sendma(driver,mail):
    with open('proxyxoay.txt', 'r') as file:
        lines = file.readlines()
    # Choose a random line from the list of lines
    random_line = random.choice(lines)

    # Print the randomly chosen line
    print(random_line)
    proxies = {
        'https': 'http://192.168.1.9:' + str(random_line.strip())
    }
    ten = ["Tran", "Nguyen", "Le", "Pham", "Huynh", "Hoang", "Phan", "Vu", "Dang", "Bui", "Do", "Ho", "Ngo", "Duong", "Ly", "Dao", "Dinh", "Doan", "Vuong", "Trinh"]
    tenmail = random.choice(ten)
    tenmien = "@navalcadets.com"
    
    
    url = "https://api.mail.tm/token"
    payload = {
        "address": mail,
        "password": "thongvo123"
    }

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    response = requests.post(url, headers=headers, json=payload,proxies=proxies)
    print(response.json())
    




    url = "https://api.mail.tm/token"

    payload = {
        "address":  mail,
        "password": "thongvo123"
    }

    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json"
    }

    response = requests.post(url, headers=headers, json=payload,proxies=proxies)
    print(response.json())
    


    thu = response.json()['token']
    print(thu)
    header = {"authorization": "Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiJ9.eyJpYXQiOjE2Njk5Njc3ODMsInJvbGVzIjpbIlJPTEVfVVNFUiJdLCJ1c2VybmFtZSI6Im5wbWtjbGlqZEBrYXJlbmtleS5jb20iLCJpZCI6IjYzODlhZmE3YTY5ZDBmZmMyMTE0MzIxYyIsIm1lcmN1cmUiOnsic3Vic2NyaWJlIjpbIi9hY2NvdW50cy82Mzg5YWZhN2E2OWQwZmZjMjExNDMyMWMiXX19.-odZWTOPkc1txvfqWXG3GoBAyWwI9q_copAGufZFcC1qTWR4gEiuDwJiZEwgOfpS0uBaaiWrsCTC4ubS9a3_5g"}
    header["authorization"] = "Bearer " + thu
    print(header)
    r = requests.get('https://api.mail.tm/messages', headers=header,proxies=proxies)
    mail_katusu = r.text
    mail_katusu = json.loads(mail_katusu)
    if str(mail_katusu["hydra:member"]) != "[]":
        id = mail_katusu["hydra:member"][0]["id"]
        r = requests.get(f'https://api.mail.tm/messages/{id}', headers=header,proxies=proxies)
        mail = r.text
        mail = json.loads(mail)
        text = mail["text"]
        print(text)
        numbers = re.findall(r'\d{6}', text)
        print(numbers[-1])
        i = 0
        for numbers1 in numbers[-1]:
            driver.implicitly_wait(500)
            driver.find_element(By.XPATH,f"//input[@inputmode='numeric' and @data-id='{i}']").send_keys(int(numbers1))
            i+=1
    time.sleep(1)
    
def okx(driver,keyokx):
    try:
        close_all_tabs_except_main(driver)
    except:
        print("ok")
    words = keyokx.split()

    window_handles = driver.window_handles
    driver.switch_to.window(window_handles[0])
    driver.get("chrome-extension://mcohilncbfahbmgdjkbpemcciiolgcge/popup.html")
    #time.sleep(100000)
    try:
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,'//input[@placeholder="Enter your password"]').send_keys("A12345678a")
        driver.implicitly_wait(2)
        driver.find_element(By.CLASS_NAME,"btn-content").click()
    except:
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//span[text()='Import wallet']").click()
        time.sleep(2)
        try:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,"//div[contains(text(), 'Seed phrase or private key')]").click()
        except:
            driver.implicitly_wait(10)
            driver.find_element(By.XPATH,"//div[@title='Import wallet'][text()='Import wallet']").click()
        time.sleep(5)
        for i in range(1, 13):
            #input_index = 3  # Giả sử bạn muốn thao tác với phần tử thứ 3
            input_element = driver.find_element(By.XPATH,f"//div[@class='mnemonic-words-inputs__container']/div[{i}]//input")
            input_element.send_keys(str(words[i - 1]))
            input_element.send_keys(Keys.ENTER)

            
        time.sleep(5)
        js_code = """
        var xpath = "//button[@type='submit']//span[text()='Confirm']";
        var elements = document.evaluate(xpath, document, null, XPathResult.ORDERED_NODE_SNAPSHOT_TYPE, null);
        if (elements.snapshotLength > 0) {
            elements.snapshotItem(0).click();
        }
        """

        # Thực hiện mã JavaScript trên trang
        driver.execute_script(js_code)
        
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//div[text()='Password']").click()
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH, "//button[.//span[text()='Next']]").click()

        
        time.sleep(2)
        time.sleep(1)
        driver.implicitly_wait(10)
        input_elements = driver.find_elements(By.CSS_SELECTOR, 'input.okui-input-input[type="password"]')

        # Gửi giá trị vào từng phần tử input
        for input_element in input_elements:
            input_element.send_keys("A12345678a")
        time.sleep(1)
        time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//button[@type='submit']//span[text()='Confirm']").click()
        timeout = 30
        start_time = time.time()

        # Loop to check the page source
        while True:
            if "Start your Web3 journey" in driver.page_source:
                break
            elif time.time() - start_time > timeout:
                print("Timeout: Text not found within 30 seconds.")
                driver.quit()
                break
            else:
                print("chưa xong")
                time.sleep(1)  # Optional: Add a short sleep to reduce CPU usage
        print("xong")    
def start(tk, mk, fa, counter, profile, diachivi, keyokx,keyok1222):
    
    usename = tk
    pass123 = mk
    # usename = tk
    # pass123 = mk
    tai = 0
    counter = counter + 1
    text = counter
    # with open('proxyxoay.txt', 'r') as file:
    #         lines = file.readlines()  # Đọc tất cả dòng và lưu vào danh sách lines
    #         if lines:  # Kiểm tra xem danh sách có dòng nào khôngs
    #             if 1 <= text <= len(lines):
    #                 line_to_print = lines[text - 1].strip()  # Lấy dòng tương ứng với chỉ số và loại bỏ khoảng trắng
    #                 print(line_to_print)
    #                 proxy191 = line_to_print
    #                 ip12,port12 = proxy191.split(":")
    #             else:
    #                 print("Chỉ số dòng không hợp lệ")
    #         else:
    #             print("Tệp tin trống")  # In ra thông báo nếu tệp tin không có nội dung
    # while True:
    #     url = f"http://192.168.1.11:22222//api/v1/status?proxy={str(port12)}"

    #     # Gửi yêu cầu GET tới URL
    #     response = requests.get(url)

    #     # Kiểm tra nếu yêu cầu thành công
    #     if response.status_code == 200:
    #         data = response.json()  # Chuyển đổi phản hồi JSON
    #         if data.get("data") and data["data"].get("status") == "CONNECTED":
    #             print("Trạng thái kết nối là 'CONNECTED'.")
    #             break  # Thoát vòng lặp khi trạng thái là CONNECTED
    #         else:
    #             print("Trạng thái không phải 'CONNECTED'. Đợi 10 giây trước khi thử lại.")
    #             time.sleep(10)
    #     else:
    #         print(f"Không thể truy cập API. Mã lỗi: {response.status_code}")
    # tk = tk.split(":")
    print("thanh")
    options = Options()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-infobars")
    options.add_argument("--start-maximized")

    characters = string.ascii_letters + string.digits

    # Tạo tên thư mục ngẫu nhiên với 7 ký tự
    random_folder_name = ''.join(random.choice(characters) for _ in range(7))
    with open('profiletamthoi.txt', 'r') as file:
        profile = file.read().strip()
    temp_dir = os.path.join(profile, random_folder_name)
    options.add_argument('--user-data-dir=' + temp_dir +  "/Data/profile")
    with open('metamaks.txt', 'r') as file:
        metamaks = file.read().strip()
    with open('proxy1.txt', 'r') as file:
        proxy1 = file.read().strip()
    with open('funcapcha.txt', 'r') as file:
        funcapcha = file.read().strip()
    options.add_argument(f'--load-extension={metamaks},{proxy1},{funcapcha}')
    options.add_argument("--disable-gpu")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36"
    )
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--lang=en")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(f"--remote-debugging-port={random.randint(1000,9999)}")  # Dùng CDP nếu cần
    chromedriver_path = "chromedriver.exe" 
    #driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options1)
    service = ChromeService(executable_path=chromedriver_path)
    driver = webdriver.Chrome(service=service, options=options)
    # driver = uc.Chrome(options=options, keep_alive=True,version_main=108)
    value_spinBox_2 = 501
    value_spinBox_3 = 800
    windows_info = {}
    text = counter
    so = 10
    
    if value_spinBox_2 > value_spinBox_3:
        toado = value_spinBox_2 - value_spinBox_3
    if value_spinBox_2 < value_spinBox_3:
        toado = value_spinBox_3 - value_spinBox_2
    toado = 100
    if value_spinBox_2 == 500 or value_spinBox_3 == 500:
        row = 0
        col = 0
        if value_spinBox_2 < 500 or value_spinBox_3 < 500:
            # value_spinBox_2 = 500
            # value_spinBox_3 = 500
            driver.set_window_size(500, 500)
        else:
            print("thì thôi")
            chieudai = random.randint(1,50)
            chieurong = random.randint(1,50)
            driver.set_window_size(value_spinBox_2 + chieudai, value_spinBox_3 + chieurong)
        for i in range(text):
            if toado == 100:
                toado1 = 400
            elif toado == 200:
                toado1 = 300
            elif toado == 300:
                toado1 = 200
            elif toado == 400:
                toado1 = 100
            position_x = toado1 * col
            position_y = (value_spinBox_3 - toado) * row
            print(abs(toado1 - 500))
            
            driver.set_window_position(position_x , position_y)
            windows_info[driver.current_window_handle] = (position_x, position_y)
            
            col += 1
            if value_spinBox_2 == 500 and value_spinBox_3 == 500:
                caccot = so
            else:
                if value_spinBox_2 == 500 and value_spinBox_3 == 300:
                    caccot = 20
                elif value_spinBox_2 == 500 and value_spinBox_3 == 200:
                    caccot = 25
                elif value_spinBox_2 == 500 and value_spinBox_3 == 100:
                    caccot = 30
                elif value_spinBox_2 < 500:
                    tong = 500 * so
                    caccot = tong / toado1
                elif value_spinBox_3 < 500:
                    tong = 500 * so
                    caccot = tong / toado1
            print(caccot) 
            if col >= int(caccot):  # Assuming you want 7 columns
                col = 0
                row += 1
    else:
        driver.set_window_size(value_spinBox_2, value_spinBox_3)
        position_x = 50 * ((text - 1) % so)
        position_y = 200 * ((text - 1) // so)
        driver.set_window_position(position_x, position_y)
    for window_handle, (x, y) in windows_info.items():
        driver.switch_to.window(window_handle)
        driver.set_window_position(x, y)
    driver.set_window_size(value_spinBox_2, value_spinBox_3)
    time.sleep(2)
    position_x = 0
    driver.get("chrome://version/")
    time.sleep(1)
    driver.implicitly_wait(10)
    link332de2 = driver.find_element(By.ID,"profile_path").text
    profile_path = link332de2
    profile_path = base_path = profile_path.rsplit('\\', 3)[0]
    print(link332de2)
    NOPECHA_KEY = "sub_1QArc4CRwBwvt6pt8wTyBELu"
    load_extension_page(driver,"chrome-extension://hlifkpholllijblknnmbfagnkjneagid/popup/popup.html")
    #driver.get("https://www.google.com/")
    try:
        main_tab = driver.current_window_handle

        # Thời gian chờ tối đa (tổng thời gian theo dõi)
        timeout = 10
        start_time = time.time()

        # Theo dõi và đóng tab không mong muốn
        while time.time() - start_time < timeout:
            # Lặp qua tất cả các tab đang mở
            for handle in driver.window_handles:
                if handle != main_tab:  # Nếu không phải tab chính
                    driver.switch_to.window(handle)
                    driver.close()  # Đóng tab không mong muốn
                    print("Đã đóng một tab không mong muốn!")

            # Quay lại tab chính
            driver.switch_to.window(main_tab)

            # Nếu không còn tab nào lạ, thoát vòng lặp
            if len(driver.window_handles) == 1:
                break

            # Chờ một chút trước khi kiểm tra lại
            time.sleep(0.5)

        print("Tất cả các tab không mong muốn đã được đóng!")
        driver.get("chrome://extensions/")
        driver.implicitly_wait(10)
        #time.sleep(1)
        driver.execute_script(
            "chrome.management.setEnabled('mnloefcpaepkpmhaoipjkpikbnkmbnic', true, function() {"
            "    if (chrome.runtime.lastError) {"
            "        console.error(chrome.runtime.lastError.message);"
            "    } else {"
            "        console.log('Extension disabled successfully');"
            "    }"
            "});"
        )
        #time.sleep(1)
        with open('proxyxoay1.txt', 'r') as file:
            lines = file.readlines()  # Đọc tất cả dòng và lưu vào danh sách lines
            if lines:  # Kiểm tra xem danh sách có dòng nào khôngs
                if 1 <= text <= len(lines):
                    line_to_print = lines[text - 1].strip()  # Lấy dòng tương ứng với chỉ số và loại bỏ khoảng trắng
                    print(line_to_print)
                    proxy191 = line_to_print
                    ip12,port12 = proxy191.split(":")
                else:
                    print("Chỉ số dòng không hợp lệ")
            else:
                print("Tệp tin trống")  # In ra thông báo nếu tệp tin không có nội dung
        driver.get("chrome-extension://mnloefcpaepkpmhaoipjkpikbnkmbnic/options.html")
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[1]").clear()
        driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[1]").send_keys(str(ip12))
        driver.implicitly_wait(10)
        
        driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[2]").clear()
        driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[2]").send_keys(str(port12))
        driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[2]").send_keys(Keys.RETURN)
        driver.refresh()
        # time.sleep(1)
        # driver.get("chrome-extension://mnloefcpaepkpmhaoipjkpikbnkmbnic/options.html")
        # time.sleep(1)
        # driver.implicitly_wait(10)
        # print(tk)
        # #time.sleep(10000)
        
        # ip = tk[0]
        # port = tk[1]
        # driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[1]").clear()
        # driver.implicitly_wait(10)
        # driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[1]").send_keys("192.168.1.11")
        # driver.implicitly_wait(10)
        # with open('proxyxoay.txt', 'r') as file:
        #     lines = file.readlines()  # Đọc tất cả dòng và lưu vào danh sách lines
        #     if lines:  # Kiểm tra xem danh sách có dòng nào không
        #         random_line = random.choice(lines)  # Chọn ngẫu nhiên một dòng từ danh sách
        #         print(random_line.strip())
        #         proxy191 = random_line.strip()  # In ra dòng đã chọn sau khi loại bỏ khoảng trắng (nếu có)
        #     else:
        #         print("Tệp tin trống")  # In ra thông báo nếu tệp tin không có nội dung
        # driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[2]").clear()
        # driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[2]").send_keys(str(proxy191))
        # driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[2]").send_keys(Keys.RETURN)
        # driver.refresh()
        # time.sleep(1)
        # driver.refresh()
        # time.sleep(1)
        # driver.refresh()
        # time.sleep(2)
        # driver.get("chrome-extension://mnloefcpaepkpmhaoipjkpikbnkmbnic/popup.html")
        # time.sleep(1)
        # driver.implicitly_wait(10)
        # driver.find_element(By.XPATH,"//span[text()='HTTP PROXY']").click()
        # time.sleep(2)
        # driver.get("https://www.google.com/")
        #time.sleep(5)
        # driver.get("chrome://extensions/")
        # time.sleep(10000000)
        # driver.implicitly_wait(10)
        # #time.sleep(1)
        # driver.execute_script(
        #     "chrome.management.setEnabled('mnloefcpaepkpmhaoipjkpikbnkmbnic', true, function() {"
        #     "    if (chrome.runtime.lastError) {"
        #     "        console.error(chrome.runtime.lastError.message);"
        #     "    } else {"
        #     "        console.log('Extension disabled successfully');"
        #     "    }"
        #     "});"
        # )
        # #time.sleep(1)
        # driver.get("chrome-extension://mnloefcpaepkpmhaoipjkpikbnkmbnic/options.html")
        # with open('proxyxoay.txt', 'r') as file:
        #     lines = file.readlines()  # Đọc tất cả dòng và lưu vào danh sách lines
        #     if lines:  # Kiểm tra xem danh sách có dòng nào khôngs
        #         if 1 <= text <= len(lines):
        #             line_to_print = lines[text - 1].strip()  # Lấy dòng tương ứng với chỉ số và loại bỏ khoảng trắng
        #             print(line_to_print)
        #             proxy191 = line_to_print
        #             # ip,port,tk,mk = proxy191.split(":")
        #         else:
        #             print("Chỉ số dòng không hợp lệ")
        #     else:
        #         print("Tệp tin trống")  # In ra thông báo nếu tệp tin không có nội dung
        # driver.implicitly_wait(10)
        # driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[1]").clear()
        # driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[1]").send_keys("192.168.1.11")
        # driver.implicitly_wait(10)
        
        # driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[2]").clear()
        # driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[2]").send_keys(str(proxy191))
        # driver.find_element(By.XPATH,"/html/body/div/div[2]/div[1]/div/section[1]/div[1]/input[2]").send_keys(Keys.RETURN)
        # driver.refresh()
        # time.sleep(1)
        # time.sleep(1)
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"/html/body/div/div[1]/ul[1]/li[4]/a").click()
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"username").clear()
        driver.find_element(By.ID,"username").send_keys("vovantai")
        driver.implicitly_wait(10)
        driver.find_element(By.ID,"password").clear()
        driver.find_element(By.ID,"password").send_keys("thongvo123")
        driver.find_element(By.ID,"password").send_keys(Keys.RETURN)
        # time.sleep(1)
        # driver.refresh()
        
        # driver.get("chrome-extension://mnloefcpaepkpmhaoipjkpikbnkmbnic/popup.html")
        # driver.implicitly_wait(10)
        # driver.find_element(By.ID,"direct").click()
        #driver.find_element(By.XPATH,"/html/body/ul/li[4]").click()
        close_all_tabs_except_main(driver)
        okx(driver,keyokx)
        
        # time.sleep(1)
        # #driver.find_element(By.XPATH,"/html/body/ul/li[4]").click()
        # time.sleep(1)

        # link123123 = None
        # letters_and_digits = string.ascii_letters + string.digits
        
        # # logmetmaks(driver,keyokx)
        # # time.sleep(100000)
        # window_handles = driver.window_handles
        # driver.switch_to.window(window_handles[0])
        # try:
        #     close_all_tabs_except_main(driver)
        # except:
        #     print("ok")
        # load_extension_page(driver,"chrome-extension://hlifkpholllijblknnmbfagnkjneagid/popup/popup.html#/")
        # time.sleep(2)
        driver.get("chrome-extension://mnloefcpaepkpmhaoipjkpikbnkmbnic/popup.html")
        driver.implicitly_wait(10)
        #driver.find_element(By.ID,"direct").click()
        driver.find_element(By.XPATH,"/html/body/ul/li[4]").click()
        time.sleep(5)
        close_all_tabs_except_main(driver)
        #time.sleep(1000000)
        #driver.set_window_size(501, 501)
        
        # usename,pass123 = dangkitwiter(driver,link332de2)
        # time.sleep(10000000)
        # login_twitter(driver,tk,mk,fa)
        # link = driver.current_url
        # if link == "https://x.com/account/access":
        #     print("out")
        # else:
        #     load_extension_page(driver,"https://x.com/intent/retweet?tweet_id=1850139232393892341")
        #     repost(driver)
        #     load_extension_page(driver,"https://x.com/intent/retweet?tweet_id=1854859578690937160")
        #     repost(driver)
            #time.sleep(1000000)
            # load_extension_page(driver,"https://x.com/intent/follow?screen_name=DavidDobrik")
            # folow(driver,tk,mk,fa)
            # load_extension_page(driver,"https://x.com")
            # time.sleep(5)
        #driver.set_window_size(501, 501)
        time.sleep(5)
        load_extension_page(driver,"https://quest.intract.io/")
        driver.implicitly_wait(600)
        #7458
        #driver.find_element(By.XPATH,"//span[contains(text(), 'Sign In') and contains(@class, 'button')]")
        driver.find_element(By.XPATH,"//div[contains(text(), 'Sign In') and contains(@class, '_primary_button_1sypl_1')]")
        #driver.find_element(By.XPATH,"//span[contains(@class, 'button') and contains(@class, 'btn-sm') and contains(@class, 'navbar-button') and text()='Sign In']")
        # js_code = """
        # const xpath = "//span[contains(@class, 'button') and contains(@class, 'btn-sm') and contains(@class, 'navbar-button') and text()='Sign In']";
        # const element = document.evaluate(xpath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
        # if (element) {
        #     element.click();
        # } else {
        #     console.error('Element not found');
        # }
        # """
        # js_script = '''
        # var signInButton = document.evaluate("//div[contains(text(), 'Sign In')]", document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
        # if (signInButton) {
        #     signInButton.click();
        # } else {
        #     console.log('Sign In button not found');
        # }
        # '''

        # # Execute the JavaScript code using Selenium
        # # Execute the JavaScript using Selenium
        # driver.execute_script(js_script)
        #7458
        # script = "var element = document.evaluate(\"//span[contains(text(), 'Sign In') and contains(@class, 'button')]\"," \
        #         " document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue; " \
        #         "if (element) { element.click(); } else { console.log('Element not found'); }"

        # # Thực thi đoạn mã JavaScript
        # driver.execute_script(script)
        script = """
            const xpath = "//div[contains(text(), 'Sign In') and contains(@class, '_primary_button_1sypl_1')]";
            const element = document.evaluate(xpath, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;
            if (element) {
                element.click();
            } else {
                console.log('Element not found');
            }
        """

        # Thực thi đoạn mã JavaScript
        driver.execute_script(script)
        driver.implicitly_wait(600)
        driver.find_element(By.XPATH,'//article[h5[text()="OKX Wallet"]]')
        script = '''
        var metaMaskButton = document.evaluate(
        '//article[h5[text()="OKX Wallet"]]',
        document,
        null,
        XPathResult.FIRST_ORDERED_NODE_TYPE,
        null
        ).singleNodeValue;

        if (metaMaskButton) {
        metaMaskButton.click();
        } else {
        console.log("MetaMask button not found.");
        }
        '''

        # Execute the script using Selenium
        driver.execute_script(script)
        time.sleep(5)
        # Lấy danh sách các tab (window handles)
        tabs = driver.window_handles

        # Chuyển sang tab đầu tiên (index 0)
        driver.switch_to.window(tabs[-1])
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//button[span/div[text()='Connect']]").click()
        time.sleep(5)
        tabs = driver.window_handles

        # Chuyển sang tab đầu tiên (index 0)
        driver.switch_to.window(tabs[0])
        time.sleep(5)
        # Lấy danh sách các tab (window handles)
        tabs = driver.window_handles

        # Chuyển sang tab đầu tiên (index 0)
        driver.switch_to.window(tabs[-1])
        driver.implicitly_wait(10)
        driver.find_element(By.XPATH,"//button[span/div[text()='Confirm']]").click()
        time.sleep(10)
        #time.sleep(10)
        time.sleep(5)
        # Lấy danh sách các tab (window handles)
        tabs = driver.window_handles

        # Chuyển sang tab đầu tiên (index 0)
        driver.switch_to.window(tabs[0])
        lan = 0
        solan = 0
        while True:
            load_extension_page(driver,"https://quest.intract.io/profile")
            time.sleep(5)
            driver.implicitly_wait(60)
            driver.find_element(By.XPATH,'//div[contains(@class, "button") and contains(@class, "white") and contains(text(), "Edit Profile")]')
            script = '''
            var editProfileButton = document.evaluate(
            '//div[contains(@class, "button") and contains(@class, "white") and contains(text(), "Edit Profile")]',
            document,
            null,
            XPathResult.FIRST_ORDERED_NODE_TYPE,
            null
            ).singleNodeValue;

            if (editProfileButton) {
            editProfileButton.click();
            } else {
            console.log("Edit Profile button not found.");
            }
            '''

            # Execute the script using Selenium
            driver.execute_script(script)
            time.sleep(2)
            driver.implicitly_wait(10)
            twiter111 = driver.find_element(By.XPATH,"//button[.//img[@src='/assets/img/brands/twitter.png']]").text
            if "Twitter" not in twiter111:
                print("Không tìm thấy 'Twitter' trong văn bản.")
                break
            else:

                print('The element with the specified classes does not exist.')
                                
                
                load_extension_page(driver,"https://quest.intract.io/profile")
                driver.implicitly_wait(60)
                driver.find_element(By.XPATH,'//div[contains(@class, "button") and contains(@class, "white") and contains(text(), "Edit Profile")]')
                script = '''
                var editProfileButton = document.evaluate(
                '//div[contains(@class, "button") and contains(@class, "white") and contains(text(), "Edit Profile")]',
                document,
                null,
                XPathResult.FIRST_ORDERED_NODE_TYPE,
                null
                ).singleNodeValue;

                if (editProfileButton) {
                editProfileButton.click();
                } else {
                console.log("Edit Profile button not found.");
                }
                '''

                # Execute the script using Selenium
                driver.execute_script(script)
                time.sleep(2)
                #driver.implicitly_wait(10)
                script = '''
                var twitterButton = document.evaluate(
                '//button[contains(@class, "button") and contains(@class, "w-button") and contains(.,"Twitter")]',
                document,
                null,
                XPathResult.FIRST_ORDERED_NODE_TYPE,
                null
                ).singleNodeValue;

                if (twitterButton) {
                twitterButton.click();
                } else {
                console.log("Twitter button not found.");
                }
                '''

                # Run the script using execute_script
                driver.execute_script(script)
                while True:
                    if "Connect your Twitter account" in driver.page_source:
                        break
                tabs = driver.window_handles

                # Loop through all tabs
                for tab in tabs:
                    # Switch to each tab
                    driver.switch_to.window(tab)
                    
                    # Get the current URL
                    current_url = driver.current_url
                    
                    # Check if the current tab contains the desired URL
                    if "https://twitter.com/i/oauth2/authorize" in current_url:
                        print(f"Found the desired tab: {current_url}")
                        break
                if lan == 0:
                    driver.implicitly_wait(600)
                    driver.find_element(By.XPATH,"//input[@name='text' and @autocomplete='username']").send_keys(usename)
                    driver.implicitly_wait(10)
                    driver.find_element(By.XPATH,"//span[text()='Next']").click()
                    driver.implicitly_wait(10)
                    driver.find_element(By.NAME,"password").send_keys(pass123)
                    driver.implicitly_wait(10)
                    driver.find_element(By.XPATH,"//button[.//span[text()='Log in']]").click()
                    # if fa is None:
                    #     print("không cần")
                    # else:
                    #     driver.implicitly_wait(10)
                    #     token2929 = driver.find_element(By.NAME,"text")

                    #     print(fa)
                    #     url = "https://2fa.live/tok/" + fa
                    #     response = requests.get(url)
                    #     data = json.loads(response.text)
                    #     token = data["token"]

                    #     print(token)

                    #     token2929 = driver.find_element(By.NAME,"text")
                    #     actions = ActionChains(driver)
                    #     actions.send_keys_to_element(token2929, token).perform()
                    #     driver.implicitly_wait(10)
                    #     next_button = driver.find_element(By.XPATH,"//span[text()='Next']")
                    #     actions.move_to_element(next_button)
                    #     actions.click().perform()
                    #     time.sleep(0.1)
                    #     try:
                    #         driver.find_element(By.XPATH,"//span[text()='Next']")
                    #         driver.quit()
                    #     except NoSuchElementException:
                    #         print("tiep")
                    #     time.sleep(1)
                time.sleep(5)
                
                driver.implicitly_wait(30)
                driver.find_element(By.XPATH,"//button[.//span[text()='Authorize app']]")
                script = """
                // Locate the "Authorize app" button using XPath
                var authorizeButton = document.evaluate("//button[.//span[text()='Authorize app']]", document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;

                // Check if the element exists before trying to click
                if (authorizeButton) {
                    // Trigger a click on the button
                    authorizeButton.click();
                } else {
                    console.log("Authorize app button not found!");
                }
                """

                # Execute the script
                driver.execute_script(script)
                print("thực hiện xong ")
                
                
                try:
                    driver.switch_to.window(driver.window_handles[0])
                except:
                    print("ok")
                time.sleep(10)
                load_extension_page(driver,"https://quest.intract.io/")
                time.sleep(2)
                solan +=1
                lan+=1


        
        
        if driver:
            driver.quit()
        if os.path.exists(profile_path):
            try:
                # Xóa thư mục profile
                shutil.rmtree(profile_path)
                print(f"Profile đã được xóa: {profile_path}")
            except Exception as e:
                print(f"Không thể xóa profile: {e}")
        else:
            print(f"Thư mục profile không tồn tại: {profile_path}")
        #time.sleep(10000)

    except Exception as e:
        print(e)
        time.sleep(1)
        filename = "tkloi.txt"
        if not os.path.isfile(filename):
            with open(filename, "w", encoding="utf8") as f:
                f.write(diachivi)  # Write the content of link332de2
        else:
            with open(filename, "a", encoding="utf8") as f:
                f.write("\n" + diachivi)  # Append the content of link332de2
        # In ra chi tiết của ngoại lệ
        print("Đã xảy ra một lỗi:", e)
        #time.sleep(`100000`)
        print("lỗi ở hồ sơ",text)
        time.sleep(1)
        if driver:
            driver.quit()
        if os.path.exists(profile_path):
            try:
                # Xóa thư mục profile
                shutil.rmtree(profile_path)
                print(f"Profile đã được xóa: {profile_path}")
            except Exception as e:
                print(f"Không thể xóa profile: {e}")
        else:
            print(f"Thư mục profile không tồn tại: {profile_path}")
    # finally:
    #     if driver:
    #         driver.quit()
    #     if os.path.exists(profile_path):
    #         try:
    #             # Xóa thư mục profile
    #             shutil.rmtree(profile_path)
    #             print(f"Profile đã được xóa: {profile_path}")
    #         except Exception as e:
    #             print(f"Không thể xóa profile: {e}")
    #     else:
    #         print(f"Thư mục profile không tồn tại: {profile_path}")
if __name__ == '__main__':
    proxies = []
    tk1 = []
    mk1 = []
    fa1 = []
    profile1 = []
    diachivimeta = []
    keyokx1 = []
    keyok12222 = []
    print("tool twitter")
    print("che do nuoi nick nhap phim 1")
    print("che do dang ki nhap phim gi cung dc")

    workbook = load_workbook(filename='data.xlsx')
    worksheet = workbook.active
    total_rows = worksheet.max_row
    current_row = 1
    num_rows = int(input('Enter the number of rows to print (0 to stop): '))

    task_queue = queue.Queue()
    counter_queue = queue.Queue()
    max_workers = num_rows

    # Khởi tạo các giá trị counter và đưa vào hàng đợi
    for i in range(max_workers):
        counter_queue.put(i)

    # Start worker threads
    threads = []
    for _ in range(max_workers):
        t = threading.Thread(target=worker, args=(task_queue, counter_queue))
        t.start()
        threads.append(t)

    count1 = 0

    while current_row <= total_rows:
        if num_rows == 0 or num_rows > (total_rows - current_row + 1):
            break

        for row in worksheet.iter_rows(min_row=current_row, max_row=current_row + num_rows - 1, values_only=True):
            tk = row[0]
            mk = row[1]
            fa = row[2]
            profile = row[3]
            diachivi = row[4]
            keyokx = row[5]
            keyok1222 = row[6]
            print(tk)
            print(mk)
            print(fa)
            profile1.append(profile)
            tk1.append(tk)
            mk1.append(mk)
            fa1.append(fa)
            keyokx1.append(keyokx)
            diachivimeta.append(diachivi)
            keyok12222.append(keyok1222)

        count1 += 1

        for i in range(len(tk1)):
            task_queue.put((tk1[i], mk1[i], fa1[i], profile1[i], diachivimeta[i], keyokx1[i],keyok12222[i]))
            time.sleep(1)

        current_row += num_rows
        proxies = []
        tk1 = []
        mk1 = []
        fa1 = []
        profile1 = []
        diachivimeta = []
        keyokx1 = []
        keyok12222 = []
    # Block until all tasks are done
    task_queue.join()

    # Stop worker threads
    for _ in range(max_workers):
        task_queue.put(None)
    for t in threads:
        t.join()

                
        
        